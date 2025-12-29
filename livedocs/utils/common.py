import os
from datetime import date, datetime, time
from typing import Any, Callable


import polars as pl
import requests
from tqdm.auto import tqdm
import base64
import decimal
import uuid
import json

from IPython.display import display

from livedocs.types import (
    DatabaseType,
    ElementDataSource,
    ElementDatasourceType,
)


def get_run_context() -> str:
    current_run_context = "edit_mode"
    match os.getenv("LIVEDOCS_RUN_CONTEXT"):
        case "logic":
            current_run_context = "edit_mode"
        case "scheduled":
            current_run_context = "scheduled_runs"
        case "webhook":
            current_run_context = "webhook_runs"
        case _:
            current_run_context = "unknown_run_context"
    return current_run_context


def middleman_debug(label: str, data=None, level: str = "info"):
    """
    Send structured debug output to Middleman logs.

    This emits a Jupyter display with metadata:
      - middleman_debug: True
      - middleman_debug_label: <label>
      - middleman_debug_level: <level>

    Middleman intercepts these displays and logs the pretty-printed payload
    (along with runtime/report context on the server side). Outside a
    Middleman runtime this is a no-op display.

    Args:
        label: A descriptive label for the debug output
        data: The data to debug (will be JSON serialized if possible)
        level: Debug level, either "error" or "info" (default: "info")
    """
    if level not in ("error", "info"):
        level = "info"

    try:
        content_str = json.dumps(data, indent=2, default=str, ensure_ascii=False)
        mime_type = "application/json"
    except Exception:
        content_str = str(data)
        mime_type = "text/plain"

    print(f"[Middleman Debug - {level.upper()}] {label} {content_str}")
    display(
        {mime_type: content_str},
        metadata={
            "middleman_debug": True,
            "middleman_debug_label": label,
            "middleman_debug_level": level.lower(),
        },
        raw=True,
    )


def serializer(obj):
    """
    Serializes an object to a JSON-compatible format.
    """

    if obj is None:
        return None
    elif isinstance(obj, bool):
        return obj
    elif isinstance(obj, (datetime, date, time)):
        return obj.isoformat()
    elif isinstance(obj, decimal.Decimal):
        return str(obj)
    elif isinstance(obj, uuid.UUID):
        return str(obj)
    elif isinstance(obj, bytes):
        return base64.b64encode(obj).decode("utf-8")
    elif isinstance(obj, dict):
        return {k: serializer(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [serializer(item) for item in obj]
    elif isinstance(obj, (set, tuple, frozenset)):
        return [serializer(item) for item in obj]
    elif isinstance(obj, complex):
        return {"real": obj.real, "imag": obj.imag}
    else:
        return str(obj)


def get_xlsx_sheet_names(local_path: str) -> list[str]:
    """
    Get sheet names from an xlsx file.

    Args:
        local_path: Path to the local xlsx file

    Returns:
        List of sheet names, or empty list if file can't be read
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(local_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception:
        return []


def get_query_for_datasource(
    datasource: ElementDataSource,
    limit: int | None,
    file_path: str | None = None,
    get_database_details: Callable[[str], tuple[object, dict[str, Any]]] | None = None,
) -> str:
    """
    Constructs the appropriate SQL query for a given datasource.

    This function handles query construction for different datasource types (database tables,
    files, dataframes) and accounts for database-specific syntax requirements.

    Args:
        datasource: The datasource configuration containing type and connection details
        limit: Optional row limit for query results. If None, no limit is applied.
        file_path: Optional file path for file datasources. When provided, uses
                   direct file querying syntax (SELECT * FROM <filepath>).
                   Supported extensions: .csv, .tsv, .txt, .gz, .parquet, .json,
                   .duckdb, .ddb, .xls, .xlsx, .sqlite
        get_database_details: Optional callable to retrieve database credentials.
                             Required for Snowflake datasources to get database name.

    Returns:
        str | None: The constructed SQL query string, or None if construction fails

    Raises:
        ValueError: If required datasource information is missing or invalid
    """
    # Determine limit clause for query. Default to 100,000 if no limit is specified.
    limit_clause = f" LIMIT {limit}" if limit is not None else " LIMIT 100000"
    supported_exts = get_duckdb_supported_file_extensions()

    match datasource["source_type"]:
        # Datasource Type: DATAFRAME
        case ElementDatasourceType.dataframe.value:
            if datasource["dataframe_info"] is None:
                raise ValueError("Dataframe info is required")
            df_name = datasource["dataframe_info"]["df_name"]
            return f"SELECT * FROM {df_name}{limit_clause};"

        # Datasource Type: DATABASE TABLE
        case ElementDatasourceType.database_table.value:
            if datasource["database_info"] is None:
                raise ValueError("Database info is required")
            if datasource["database_table_info"] is None:
                raise ValueError("Database table info is required")

            database_type = DatabaseType(datasource["database_info"]["database_type"])
            db_info = datasource["database_info"]
            table_info = datasource["database_table_info"]
            schema_name = table_info["schema_name"]
            table_name = table_info["table_name"]

            # Snowflake: Requires fetching database name from credentials
            # Format: "database"."schema"."table"
            if database_type == DatabaseType.Snowflake:
                if get_database_details is None:
                    raise ValueError(
                        "get_database_details is required for Snowflake datasources"
                    )
                try:
                    db_connector_id = db_info["database_connector_id"]
                    _, parsed_credentials = get_database_details(db_connector_id)
                    database_name = parsed_credentials.get("database")
                    if database_name is None:
                        raise ValueError(
                            "Database name not found in Snowflake credentials"
                        )
                    return f'SELECT * FROM "{database_name}"."{schema_name}"."{table_name}"{limit_clause};'
                except KeyError as e:
                    raise ValueError(f"Missing required Snowflake information: {e}")

            # BigQuery: Uses dataset.table format (no quotes)
            # Format: dataset.table
            elif database_type == DatabaseType.Bigquery:
                return f"SELECT * FROM {schema_name}.{table_name}{limit_clause};"

            # ClickHouse: Uses quoted identifiers with schema.table
            # Format: "schema"."table" or just "table" if no schema
            elif database_type == DatabaseType.Clickhouse:
                if schema_name:
                    return (
                        f'SELECT * FROM "{schema_name}"."{table_name}"{limit_clause};'
                    )
                else:
                    return f'SELECT * FROM "{table_name}"{limit_clause};'

            # PostgreSQL: Uses schema.table with quoted identifiers
            # Format: "schema"."table"
            elif database_type == DatabaseType.Postgres:
                return f'SELECT * FROM "{schema_name}"."{table_name}"{limit_clause};'

            # MotherDuck: Uses database.schema.table (three-part identifier)
            # Format: "database"."schema"."table"
            elif database_type == DatabaseType.Motherduck:
                database_name = db_info["database_name"]
                return f'SELECT * FROM "{database_name}"."{schema_name}"."{table_name}"{limit_clause};'

            # Databricks: Uses catalog.schema.table (no quotes)
            # Format: catalog.schema.table
            elif database_type == DatabaseType.Databricks:
                # Use catalog_name if available, otherwise fall back to database_name
                catalog_name = (
                    table_info.get("catalog_name") or db_info["database_name"]
                )
                return f"SELECT * FROM {catalog_name}.{schema_name}.{table_name}{limit_clause};"

            # Default: Most databases use database.schema.table with quotes
            # Format: "database"."schema"."table"
            else:
                database_name = db_info["database_name"]
                return f'SELECT * FROM "{database_name}"."{schema_name}"."{table_name}"{limit_clause};'

        # Datasource Type: FILE
        case ElementDatasourceType.file.value:
            if datasource["file_info"] is None:
                raise ValueError("File info is required")

            # Direct file path querying (preferred method)
            if file_path is not None:
                # Extract and validate file extension
                file_extension = None
                if "." in file_path:
                    file_extension = "." + file_path.rsplit(".", 1)[1].lower()

                if file_extension is None or file_extension not in supported_exts:
                    raise ValueError(
                        f"Unsupported file extension for direct querying: {file_extension or 'no extension'}. "
                        f"Supported extensions: {', '.join(sorted(supported_exts))}"
                    )

                # Escape single quotes to prevent SQL injection
                escaped_file_path = file_path.replace("'", "''")

                # Handle Excel files with sheet selection
                if file_extension == ".xlsx":
                    layer = datasource["file_info"].get("layer_name")
                    if layer:
                        return f"SELECT * FROM read_xlsx('{escaped_file_path}', sheet='{layer}', ignore_errors=true){limit_clause};"
                    return f"SELECT * FROM read_xlsx('{escaped_file_path}', ignore_errors=true){limit_clause};"

                # For all other supported file types
                return f"SELECT * FROM '{escaped_file_path}'{limit_clause};"

            # Fallback: Use file-name-based querying
            file_name = datasource["file_info"]["file_name"]

            # Extract and validate file extension
            file_extension = None
            if "." in file_name:
                file_extension = "." + file_name.rsplit(".", 1)[1].lower()

            if file_extension is None or file_extension not in supported_exts:
                raise ValueError(
                    f"Unsupported file extension for querying: {file_extension or 'no extension'}. "
                    f"Supported extensions: {', '.join(sorted(supported_exts))}"
                )

            # Use file-type-specific query methods
            file_type = datasource["file_info"]["file_type"]

            if file_type == "csv":
                return f"SELECT * FROM read_csv_auto('{file_name}'){limit_clause};"
            elif file_type == "xlsx":
                layer = datasource["file_info"].get("layer_name")
                if layer:
                    return f"SELECT * FROM read_xlsx('{file_name}', sheet='{layer}', ignore_errors=true){limit_clause};"
                return f"SELECT * FROM read_xlsx('{file_name}', ignore_errors=true){limit_clause};"
            else:
                # For other file types, use direct querying
                escaped_file_name = file_name.replace("'", "''")
                return f"SELECT * FROM '{escaped_file_name}'{limit_clause};"

        case _:
            raise ValueError(
                f"Unsupported datasource type: {datasource['source_type']}"
            )


def _get_dataframe_schema(df: pl.DataFrame) -> dict[str, str]:
    date_formats = [
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%B %d, %Y",
        "%d %b %Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y/%m/%d %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
    ]

    def is_date(value: str) -> bool:
        for fmt in date_formats:
            try:
                datetime.strptime(value, fmt)
                return True
            except ValueError:
                continue
        return False

    def is_number(value: str) -> bool:
        try:
            float(value)
            return True
        except ValueError:
            return False

    def map_column_type(series: pl.Series) -> str:
        # Drop nulls
        non_empty_series = series.drop_nulls()

        # Only filter out empty strings if the series is of string type
        if series.dtype == pl.Utf8:
            non_empty_series = non_empty_series.filter(non_empty_series != "")

        sample_values = non_empty_series.to_list()

        if not sample_values:
            return "STRING"

        # Check if all non-empty sample values can be numbers
        if all(
            isinstance(val, (int, float)) or (isinstance(val, str) and is_number(val))
            for val in sample_values
        ):
            return "NUMBER"

        # Check if all non-empty sample values can be dates
        if all(
            isinstance(val, datetime) or (isinstance(val, str) and is_date(val))
            for val in sample_values
        ):
            return "DATE"

        # Check the series' inherent type if it's not a string series
        if series.dtype in {pl.Float32, pl.Float64, pl.Int32, pl.Int64}:
            return "NUMBER"
        elif series.dtype in {pl.Date, pl.Datetime}:
            return "DATE"
        else:
            return "STRING"

    column_types = {col: map_column_type(df[col]) for col in df.columns}

    return column_types


def _setup_dirs():
    """
    Sets up the user files directory if it doesn't exist.
    """
    user_files_dir = os.getenv("LIVEDOCS_FILES_PATH")
    if not user_files_dir:
        raise ValueError("LIVEDOCS_FILES_PATH environment variable not set")

    os.makedirs(user_files_dir, exist_ok=True)


def _get_chunk_size(file_size_bytes: int | None) -> int:
    """
    Determines a dynamic chunk size based on the file size.
    """
    MB = 1024 * 1024
    if file_size_bytes is None:
        return 128 * 1024  # Default to 128KB if size is unknown

    if file_size_bytes < 1 * MB:  # Files < 1MB
        return 64 * 1024  # 64KB chunks
    elif file_size_bytes < 10 * MB:  # Files < 10MB
        return 128 * 1024  # 128KB chunks
    elif file_size_bytes < 100 * MB:  # Files < 100MB
        return 256 * 1024  # 256KB chunks
    elif file_size_bytes < 500 * MB:  # Files < 500MB
        return 512 * 1024  # 512KB chunks
    else:  # Files >= 500MB
        return 1 * MB  # 1MB chunks


def _download_file(
    signed_url: str,
    local_path: str,
    file_description: str,
    expected_size_bytes: int | None,
):
    """
    Downloads a file from a signed URL to a local path.
    """
    os.makedirs(os.path.dirname(local_path), exist_ok=True)

    try:
        response = requests.get(signed_url, stream=True)
        response.raise_for_status()

        total_size = expected_size_bytes
        chunk_size = _get_chunk_size(total_size)

        with (
            open(local_path, "wb") as f,
            tqdm(
                desc=f"Downloading {file_description}",
                total=total_size,
                unit="B",
                unit_scale=True,
                unit_divisor=1024,
                disable=total_size is None or total_size == 0,
                leave=True,
            ) as bar,
        ):
            for chunk in response.iter_content(chunk_size=chunk_size):
                if chunk:
                    f.write(chunk)
                    bar.update(len(chunk))

    except requests.exceptions.HTTPError as e:
        if os.path.exists(local_path):
            os.remove(local_path)
        error_detail = f"HTTP status {e.response.status_code}."
        error_detail += f" Response: {e.response.text}"
        raise RuntimeError(
            f"Failed to download {file_description}. {error_detail}"
        ) from e
    except requests.exceptions.RequestException as e:
        if os.path.exists(local_path):
            os.remove(local_path)
        raise RuntimeError(
            f"Network error while downloading {file_description}: {e}"
        ) from e
    except Exception as e:
        if os.path.exists(local_path):
            os.remove(local_path)
        raise RuntimeError(
            f"An error occurred during download of {file_description}: {e}"
        ) from e


def get_duckdb_supported_file_extensions() -> set[str]:
    """
    Returns the set of file extensions DuckDB can query directly via read_* functions.
    """
    return {
        ".csv",
        ".tsv",
        ".gz",
        ".parquet",
        ".json",
        ".duckdb",
        ".ddb",
        ".xls",
        ".xlsx",
        ".sqlite",
    }
