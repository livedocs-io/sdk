import json
import os
import shutil
import tempfile
import unittest
from datetime import datetime, timezone, timedelta

import polars as pl

from livedocs import Livedocs, LivedocsConfig
from livedocs.manager.credentials import StaticCredentialStore
from livedocs.types import Credentials
from livedocs.utils.lib.cache import QueryCache


class TestExportTable(unittest.TestCase):
    """Test export_table with dataframe datasources (no external DB needed)."""

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        os.environ["LIVEDOCS_FILES_PATH"] = self.temp_dir

        self.credentials_bundle = Credentials(
            workspace_id="ws-id",
            workspace_secrets={},
            databases={},
            built_in_vars={},
            s3_connectors={},
            google_drive_connectors={},
        )
        self.config = LivedocsConfig(
            credential_store_factory=lambda *_: StaticCredentialStore(
                self.credentials_bundle
            ),
            query_cache_factory=lambda report_id, token: QueryCache(report_id, token),
        )
        self.livedocs = Livedocs(config=self.config)
        self.livedocs.initialize("report-id", "session-token")

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)
        # Clean up export dir
        export_dir = "/tmp/livedocs_exports"
        if os.path.exists(export_dir):
            shutil.rmtree(export_dir, ignore_errors=True)

    def _datasource_json(self, df_name: str = "test_df") -> str:
        return json.dumps(
            {
                "source_type": "dataframe",
                "dataframe_info": {
                    "df_name": df_name,
                    "df_element_id": "dummy",
                },
            }
        )

    def _file_datasource_json(self, file_name: str) -> str:
        return json.dumps(
            {
                "source_type": "file",
                "database_info": None,
                "database_table_info": None,
                "dataframe_info": None,
                "file_info": {
                    "file_name": file_name,
                    "file_id": f"/{file_name}",
                    "file_has_layers": False,
                    "file_type": file_name.rsplit(".", 1)[-1],
                    "layer_name": None,
                    "connector_info": {
                        "connector_id": "",
                        "connector_name": "runtime",
                        "connector_type": "runtime",
                    },
                },
            }
        )

    def test_export_csv_basic(self):
        """Export a simple DataFrame to CSV."""
        df = pl.DataFrame({"id": [1, 2, 3], "name": ["a", "b", "c"]})
        filepath = self.livedocs.export_table(
            self._datasource_json(), format="csv", dataframe=df
        )

        self.assertTrue(filepath.endswith(".csv"))
        self.assertTrue(os.path.exists(filepath))

        result = pl.read_csv(filepath)
        self.assertEqual(result.height, 3)
        self.assertEqual(result.columns, ["id", "name"])

    def test_export_xlsx_basic(self):
        """Export a simple DataFrame to Excel."""
        df = pl.DataFrame({"id": [1, 2, 3], "value": [10.5, 20.0, 30.75]})
        filepath = self.livedocs.export_table(
            self._datasource_json(), format="xlsx", dataframe=df
        )

        self.assertTrue(filepath.endswith(".xlsx"))
        self.assertTrue(os.path.exists(filepath))

        # Read back with openpyxl to verify
        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        self.assertEqual(ws.max_row, 4)  # 1 header + 3 data rows
        self.assertEqual(ws.cell(1, 1).value, "id")
        self.assertEqual(ws.cell(2, 1).value, 1)

    def test_export_all_rows_no_pagination(self):
        """Verify export writes ALL rows, not just a page."""
        df = pl.DataFrame({"x": list(range(500))})
        filepath = self.livedocs.export_table(
            self._datasource_json(), format="csv", dataframe=df
        )

        result = pl.read_csv(filepath)
        self.assertEqual(result.height, 500)

    def test_export_csv_empty_dataframe(self):
        """Export an empty DataFrame to CSV — should produce header-only file."""
        df = pl.DataFrame({"col_a": [], "col_b": []}).cast(
            {"col_a": pl.Int64, "col_b": pl.String}
        )
        filepath = self.livedocs.export_table(
            self._datasource_json(), format="csv", dataframe=df
        )

        result = pl.read_csv(filepath)
        self.assertEqual(result.height, 0)
        self.assertEqual(result.columns, ["col_a", "col_b"])

    def test_export_xlsx_empty_dataframe(self):
        """Export an empty DataFrame to Excel."""
        df = pl.DataFrame({"col_a": [], "col_b": []}).cast(
            {"col_a": pl.Int64, "col_b": pl.String}
        )
        filepath = self.livedocs.export_table(
            self._datasource_json(), format="xlsx", dataframe=df
        )

        self.assertTrue(os.path.exists(filepath))
        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        # Header row only
        self.assertEqual(ws.cell(1, 1).value, "col_a")
        self.assertEqual(ws.max_row, 1)

    def test_export_xlsx_timezone_aware_datetimes(self):
        """Excel doesn't support tz-aware datetimes — they should be stripped."""
        df = pl.DataFrame(
            {
                "ts": [
                    datetime(2024, 1, 1, 12, 0, 0, tzinfo=timezone.utc),
                    datetime(
                        2024, 6, 15, 8, 30, 0, tzinfo=timezone(timedelta(hours=5))
                    ),
                ],
                "value": [1, 2],
            }
        )
        # Should not raise
        filepath = self.livedocs.export_table(
            self._datasource_json(), format="xlsx", dataframe=df
        )

        self.assertTrue(os.path.exists(filepath))
        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        self.assertEqual(ws.max_row, 3)  # header + 2 rows

    def test_export_csv_timezone_aware_datetimes(self):
        """CSV should handle tz-aware datetimes fine (serialised as strings)."""
        df = pl.DataFrame(
            {
                "ts": [
                    datetime(2024, 1, 1, 12, 0, 0, tzinfo=timezone.utc),
                ],
                "value": [42],
            }
        )
        filepath = self.livedocs.export_table(
            self._datasource_json(), format="csv", dataframe=df
        )
        self.assertTrue(os.path.exists(filepath))

    def test_export_csv_with_struct_columns(self):
        """Struct columns should be cast to strings for CSV."""
        df = pl.DataFrame(
            {
                "id": [1, 2],
                "meta": [{"key": "a", "val": 1}, {"key": "b", "val": 2}],
            }
        )
        self.assertTrue(isinstance(df.dtypes[1], pl.Struct))

        filepath = self.livedocs.export_table(
            self._datasource_json(), format="csv", dataframe=df
        )

        # Should not raise ComputeError
        self.assertTrue(os.path.exists(filepath))
        with open(filepath) as f:
            content = f.read()
        self.assertIn("id", content)

    def test_export_csv_with_list_columns(self):
        """List columns should be cast to strings for CSV."""
        df = pl.DataFrame({"id": [1, 2], "tags": [["x", "y"], ["z"]]})
        self.assertTrue(isinstance(df.dtypes[1], pl.List))

        filepath = self.livedocs.export_table(
            self._datasource_json(), format="csv", dataframe=df
        )
        self.assertTrue(os.path.exists(filepath))

    def test_export_xlsx_with_nested_columns(self):
        """Nested columns should be stringified for Excel too."""
        df = pl.DataFrame(
            {
                "id": [1, 2],
                "nested": [{"a": 1}, {"b": 2}],
                "list_col": [[1, 2], [3]],
            }
        )

        filepath = self.livedocs.export_table(
            self._datasource_json(), format="xlsx", dataframe=df
        )
        self.assertTrue(os.path.exists(filepath))

    def test_export_csv_with_filter_dict(self):
        """Filters passed as a dict should be applied before export."""
        df = pl.DataFrame(
            {"category": ["A", "B", "A", "C"], "amount": [10, 20, 30, 40]}
        )
        metadata = {
            "sort": None,
            "filters": [
                {
                    "column": "category",
                    "operator": "eq",
                    "value": "A",
                    "id": "f1",
                    "conjunction": "AND",
                }
            ],
            "styles": None,
            "calculations": None,
        }

        filepath = self.livedocs.export_table(
            self._datasource_json(),
            format="csv",
            dataframe=df,
            table_metadata=metadata,
        )

        result = pl.read_csv(filepath)
        self.assertEqual(result.height, 2)
        self.assertTrue(all(v == "A" for v in result["category"].to_list()))

    def test_export_csv_with_filter_json_string(self):
        """Filters passed as a JSON string (from middleman) should be parsed and applied."""
        df = pl.DataFrame(
            {"category": ["A", "B", "A", "C"], "amount": [10, 20, 30, 40]}
        )
        metadata_str = json.dumps(
            {
                "sort": None,
                "filters": [
                    {
                        "column": "category",
                        "operator": "eq",
                        "value": "A",
                        "id": "f1",
                        "conjunction": "AND",
                    }
                ],
                "styles": None,
                "calculations": None,
            }
        )

        filepath = self.livedocs.export_table(
            self._datasource_json(),
            format="csv",
            dataframe=df,
            table_metadata=metadata_str,
        )

        result = pl.read_csv(filepath)
        self.assertEqual(result.height, 2)

    def test_export_csv_with_sort(self):
        """Sort should be applied before export."""
        df = pl.DataFrame({"name": ["c", "a", "b"], "val": [3, 1, 2]})
        metadata = {
            "sort": {"column": "name", "direction": "asc"},
            "filters": None,
            "styles": None,
            "calculations": None,
        }

        filepath = self.livedocs.export_table(
            self._datasource_json(),
            format="csv",
            dataframe=df,
            table_metadata=metadata,
        )

        result = pl.read_csv(filepath)
        self.assertEqual(result["name"].to_list(), ["a", "b", "c"])

    def test_export_csv_with_sort_desc(self):
        """Descending sort should be applied."""
        df = pl.DataFrame({"val": [1, 3, 2]})
        metadata = {
            "sort": {"column": "val", "direction": "desc"},
            "filters": None,
            "styles": None,
            "calculations": None,
        }

        filepath = self.livedocs.export_table(
            self._datasource_json(),
            format="csv",
            dataframe=df,
            table_metadata=metadata,
        )

        result = pl.read_csv(filepath)
        self.assertEqual(result["val"].to_list(), [3, 2, 1])

    def test_export_with_filter_and_sort(self):
        """Both filter and sort applied."""
        df = pl.DataFrame(
            {
                "group": ["X", "Y", "X", "Y", "X"],
                "score": [50, 60, 30, 80, 10],
            }
        )
        metadata = {
            "sort": {"column": "score", "direction": "asc"},
            "filters": [
                {
                    "column": "group",
                    "operator": "eq",
                    "value": "X",
                    "id": "f1",
                    "conjunction": "AND",
                }
            ],
            "styles": None,
            "calculations": None,
        }

        filepath = self.livedocs.export_table(
            self._datasource_json(),
            format="csv",
            dataframe=df,
            table_metadata=metadata,
        )

        result = pl.read_csv(filepath)
        self.assertEqual(result.height, 3)
        self.assertEqual(result["score"].to_list(), [10, 30, 50])

    def test_export_csv_no_metadata(self):
        """table_metadata=None should export all rows unmodified."""
        df = pl.DataFrame({"a": [3, 1, 2]})
        filepath = self.livedocs.export_table(
            self._datasource_json(),
            format="csv",
            dataframe=df,
            table_metadata=None,
        )

        result = pl.read_csv(filepath)
        self.assertEqual(result["a"].to_list(), [3, 1, 2])

    def test_export_csv_mixed_types(self):
        """Handle int, float, string, bool, date, null columns."""
        df = pl.DataFrame(
            {
                "int_col": [1, 2, None],
                "float_col": [1.1, None, 3.3],
                "str_col": ["hello", "world", None],
                "bool_col": [True, False, None],
                "date_col": [
                    datetime(2024, 1, 1),
                    datetime(2024, 6, 15),
                    None,
                ],
            }
        )

        filepath = self.livedocs.export_table(
            self._datasource_json(), format="csv", dataframe=df
        )
        self.assertTrue(os.path.exists(filepath))

        result = pl.read_csv(filepath)
        self.assertEqual(result.height, 3)
        self.assertEqual(len(result.columns), 5)

    def test_export_xlsx_mixed_types(self):
        """Handle mixed types in Excel."""
        df = pl.DataFrame(
            {
                "int_col": [1, 2, None],
                "float_col": [1.1, None, 3.3],
                "str_col": ["hello", "world", None],
                "bool_col": [True, False, None],
            }
        )

        filepath = self.livedocs.export_table(
            self._datasource_json(), format="xlsx", dataframe=df
        )
        self.assertTrue(os.path.exists(filepath))

    def test_export_csv_wide_dataframe(self):
        """Export a DataFrame with many columns."""
        data = {f"col_{i}": list(range(10)) for i in range(100)}
        df = pl.DataFrame(data)

        filepath = self.livedocs.export_table(
            self._datasource_json(), format="csv", dataframe=df
        )

        result = pl.read_csv(filepath)
        self.assertEqual(len(result.columns), 100)
        self.assertEqual(result.height, 10)

    def test_export_csv_special_characters(self):
        """Strings with commas, quotes, newlines should be properly escaped."""
        df = pl.DataFrame(
            {
                "text": [
                    'has "quotes"',
                    "has, commas",
                    "has\nnewline",
                    "normal",
                ],
            }
        )

        filepath = self.livedocs.export_table(
            self._datasource_json(), format="csv", dataframe=df
        )

        result = pl.read_csv(filepath)
        self.assertEqual(result.height, 4)
        self.assertEqual(result["text"][0], 'has "quotes"')
        self.assertEqual(result["text"][1], "has, commas")

    def test_export_csv_from_file_datasource(self):
        """Export from a CSV file datasource (simulates runtime file)."""
        # Write a CSV to the files path
        csv_content = "name,age\nAlice,30\nBob,25\nCharlie,35\n"
        csv_path = os.path.join(self.temp_dir, "people.csv")
        with open(csv_path, "w") as f:
            f.write(csv_content)

        filepath = self.livedocs.export_table(
            self._file_datasource_json("people.csv"),
            format="csv",
        )

        result = pl.read_csv(filepath)
        self.assertEqual(result.height, 3)
        self.assertIn("name", result.columns)

    def test_export_xlsx_from_file_datasource(self):
        """Export a CSV file datasource to Excel format."""
        csv_content = "x,y\n1,10\n2,20\n"
        csv_path = os.path.join(self.temp_dir, "data.csv")
        with open(csv_path, "w") as f:
            f.write(csv_content)

        filepath = self.livedocs.export_table(
            self._file_datasource_json("data.csv"),
            format="xlsx",
        )

        self.assertTrue(filepath.endswith(".xlsx"))
        self.assertTrue(os.path.exists(filepath))

    def test_export_from_parquet_file(self):
        """Export from a Parquet file datasource."""
        df = pl.DataFrame({"a": [1, 2, 3], "b": ["x", "y", "z"]})
        parquet_path = os.path.join(self.temp_dir, "data.parquet")
        df.write_parquet(parquet_path)

        filepath = self.livedocs.export_table(
            self._file_datasource_json("data.parquet"),
            format="csv",
        )

        result = pl.read_csv(filepath)
        self.assertEqual(result.height, 3)

    def test_export_returns_unique_filepaths(self):
        """Each export should produce a unique file path."""
        df = pl.DataFrame({"x": [1]})
        paths = set()
        for _ in range(5):
            path = self.livedocs.export_table(
                self._datasource_json(), format="csv", dataframe=df
            )
            paths.add(path)

        self.assertEqual(len(paths), 5)

    def test_export_creates_output_directory(self):
        """Export dir /tmp/livedocs_exports should be created if missing."""
        export_dir = "/tmp/livedocs_exports"
        if os.path.exists(export_dir):
            shutil.rmtree(export_dir)

        df = pl.DataFrame({"x": [1]})
        filepath = self.livedocs.export_table(
            self._datasource_json(), format="csv", dataframe=df
        )

        self.assertTrue(os.path.exists(export_dir))
        self.assertTrue(os.path.exists(filepath))


if __name__ == "__main__":
    unittest.main()
